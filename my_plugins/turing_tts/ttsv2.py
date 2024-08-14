# Copyright 2023 LiveKit, Inc.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from __future__ import annotations

import asyncio
import base64
import dataclasses
import json
import os
from dataclasses import dataclass
import re
from typing import Any, List, Literal

import aiohttp
from livekit import rtc
from livekit.agents import tokenize, tts, utils

from .log import logger
from .models import TTSEncoding, TTSModels

_Encoding = Literal["mp3", "pcm"]


async def readline_from_stream(stream):  
    buffer = b''  
    while True:  
        # 假设我们一次从流中读取1024个字节（或更少的剩余字节）  
        chunk = await stream.read(1024)  
        if not chunk:  # 如果没有读取到任何数据，说明已经到达流的末尾  
            break  
        buffer += chunk
        # 查找换行符并处理行
        while b'\n' in buffer:  
            line, buffer = buffer.split(b'\n', 1)  # split在找到第一个换行符后停止  
            yield line.decode()  # 返回解码后的行（假设是UTF-8编码）  
    # 如果buffer中还有剩余数据（即最后一行没有换行符），则也返回它  
    if buffer:  
        yield buffer.decode()  


def _sample_rate_from_format(output_format: TTSEncoding) -> int:
    split = output_format.split("_")  # e.g: mp3_22050_32
    return int(split[1])


def _encoding_from_format(output_format: TTSEncoding) -> _Encoding:
    if output_format.startswith("mp3"):
        return "mp3"
    elif output_format.startswith("pcm"):
        return "pcm"

    raise ValueError(f"Unknown format: {output_format}")


@dataclass
class VoiceSettings:
    stability: float  # [0.0 - 1.0]
    similarity_boost: float  # [0.0 - 1.0]
    style: float | None = None  # [0.0 - 1.0]
    use_speaker_boost: bool | None = False


@dataclass
class Voice:
    id: str
    name: str
    category: str
    settings: VoiceSettings | None = None


DEFAULT_VOICE = Voice(
    id="EXAVITQu4vr4xnSDxMaL",
    name="Bella",
    category="premade",
    settings=VoiceSettings(
        stability=0.71, similarity_boost=0.5, style=0.0, use_speaker_boost=True
    ),
)

# API_BASE_URL_V1 = "https://api.elevenlabs.io/v1"
# AUTHORIZATION_HEADER = "xi-api-key"
# API_BASE_URL_V1 = "https://api.elevenlabs.io/v1"
API_BASE_URL_V1 = "http://172.16.1.88:18101/tts_audio"
# API_BASE_URL_V1 = "http://tts-v1.tuling123.com/tts/getText"
AUTHORIZATION_HEADER = "xi-api-key"


def init_emotional_tts():
    # global EmotionalTTS
    from openpyxl import load_workbook  
  
    # 加载xlsx文件  
    workbook = load_workbook(filename='/mnt1/qmw/github/agents/my_plugins/turing_tts/情感控制-JZQ-240801-1.xlsx')  
    
    # 选择工作表  
    # sheet = workbook['情绪控制']  # 根据你的工作表名称进行更改  
    sheet = workbook['Sheet8']  # 根据你的工作表名称进行更改  
    
    # 创建一个空字典来存储映射关系  
    EmotionalTTS = {}  
    
    # 遍历行（假设第一行是标题行，我们从第二行开始读取数据）  
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始读取数据  
        # row是一个元组，包含了当前行的所有值  
        # 假设A列是第一列，C列是第三列（索引从0开始）  
        a_value = row[0]  
        c_value = row[2]  
        if a_value is None or c_value is None:  
            continue
        
        # 将A列的值作为键，C列的值作为值，添加到字典中  
        EmotionalTTS[a_value] = c_value  
    return EmotionalTTS

@dataclass
class _TTSOptions:
    api_key: str
    voice: Voice
    model_id: TTSModels
    base_url: str
    encoding: TTSEncoding
    sample_rate: int
    streaming_latency: int
    word_tokenizer: tokenize.WordTokenizer
    chunk_length_schedule: list[int]

EmotionalTTS = init_emotional_tts()

class TTS(tts.TTS):
    def __init__(
        self,
        *,
        voice: Voice = DEFAULT_VOICE,
        model_id: TTSModels = "eleven_turbo_v2_5",
        api_key: str | None = None,
        base_url: str | None = None,
        encoding: TTSEncoding = "mp3_22050_32",
        streaming_latency: int = 3,
        word_tokenizer: tokenize.WordTokenizer = tokenize.basic.WordTokenizer(
            ignore_punctuation=False  # punctuation can help for intonation
        ),
        chunk_length_schedule: list[int] = [80, 120, 200, 260],  # range is [50, 500]
        http_session: aiohttp.ClientSession | None = None,
    ) -> None:
        super().__init__(
            capabilities=tts.TTSCapabilities(
                streaming=True,
            ),
            sample_rate=_sample_rate_from_format(encoding),
            num_channels=1,
        )
        api_key = api_key or os.environ.get("ELEVEN_API_KEY")
        if not api_key:
            raise ValueError("ELEVEN_API_KEY must be set")

        self._opts = _TTSOptions(
            voice=voice,
            model_id=model_id,
            api_key=api_key,
            base_url=base_url or API_BASE_URL_V1,
            encoding=encoding,
            sample_rate=self.sample_rate,
            streaming_latency=streaming_latency,
            word_tokenizer=word_tokenizer,
            chunk_length_schedule=chunk_length_schedule,
        )
        self._session = http_session

    def _ensure_session(self) -> aiohttp.ClientSession:
        if not self._session:
            self._session = utils.http_context.http_session()

        return self._session

    # async def list_voices(self) -> List[Voice]:
    #     async with self._ensure_session().get(
    #         f"{self._opts.base_url}/voices",
    #         headers={AUTHORIZATION_HEADER: self._opts.api_key},
    #     ) as resp:
    #         return _dict_to_voices_list(await resp.json())

    def synthesize(self, text: str) -> "ChunkedStream":
        return ChunkedStream(text, self._opts, self._ensure_session())

    def stream(self) -> "SynthesizeStream":
        return SynthesizeStream(self._ensure_session(), self._opts)


class ChunkedStream(tts.ChunkedStream):
    """Synthesize using the chunked api endpoint"""

    def __init__(
        self, text: str, opts: _TTSOptions, session: aiohttp.ClientSession
    ) -> None:
        super().__init__()
        self._text, self._opts, self._session = text, opts, session
        if _encoding_from_format(self._opts.encoding) == "mp3":
            self._mp3_decoder = utils.codecs.Mp3StreamDecoder()

    @utils.log_exceptions(logger=logger)
    async def _main_task(self) -> None:
        bstream = utils.audio.AudioByteStream(
            sample_rate=self._opts.sample_rate, num_channels=1
        )
        request_id = utils.shortuuid()
        segment_id = utils.shortuuid()

        voice_settings = (
            _strip_nones(dataclasses.asdict(self._opts.voice.settings))
            if self._opts.voice.settings
            else None
        )
        data = {
            "text": self._text,
            "model_id": self._opts.model_id,
            "voice_settings": voice_settings,
        }
        parameters = {
            "spk_id": "tender",
            # "instruct_text": "positive and optimistic, willing to help others.",
            # "speed":1.0,
            # "audio_sr":32000

        }
        inputJson = dict(
                text=self._text,
                type=4, # 4:演示特殊定制的一个参数
                globalId="livekit-agents-test1",
                stream=True,
                parameters=parameters
            )
        async with self._session.post(
            _synthesize_url(self._opts),
            # headers={AUTHORIZATION_HEADER: self._opts.api_key},
            json=inputJson,
        ) as resp:
            bytes_per_frame = (self._opts.sample_rate // 100) * 2
            buf = bytearray()
            resStr = ''
            async for line  in readline_from_stream(resp.content): 
                if line:
                    # result = json.loads(line.decode('utf-8'))
                    result = json.loads(line)
                    if result['code'] == 200 and result.get('index', 0) != -1:
                        # audio_data_combined = decode_speech(result['tts_speech'])
                        # torchaudio.save('output_stream.wav', audio_data_combined, 22050)

                        # audio_data = decode_speech(result['tts_speech'])
                        pcm_base64 = result['tts_speech']
                        pcmData = base64.decodebytes(pcm_base64.encode())

                        # buf.extend(pcmData)
                        for frame in bstream.write(pcmData):
                            self._event_ch.send_nowait(
                                tts.SynthesizedAudio(
                                    request_id=request_id,
                                    segment_id=segment_id,
                                    frame=frame,
                                )
                            )
                        # audio_data_list.append(audio_data)
                        print('======ttsV2=========>', result['index'])
                    elif result['index'] == -1:
                        print(result)
                        break
                    else:
                        print(f"Error: {result['message']}")
                        break
                pass

            for frame in bstream.flush():
                self._event_ch.send_nowait(
                    tts.SynthesizedAudio(
                        request_id=request_id, segment_id=segment_id, frame=frame
                    )
                )


class SynthesizeStream(tts.SynthesizeStream):
    """Streamed API using websockets"""

    def __init__(
        self,
        session: aiohttp.ClientSession,
        opts: _TTSOptions,
    ):
        super().__init__()
        self._opts, self._session = opts, session
        self._mp3_decoder = utils.codecs.Mp3StreamDecoder()

    @utils.log_exceptions(logger=logger)
    async def _main_task(self) -> None:
        self._segments_ch = utils.aio.Chan[tokenize.WordStream]()

        @utils.log_exceptions(logger=logger)
        async def _tokenize_input():
            """tokenize text from the input_ch to words"""
            word_stream = None
            async for input in self._input_ch:
                print(f'===ttsV2 _tokenize_input {input}')
                if isinstance(input, str):
                    if word_stream is None:
                        # new segment (after flush for e.g)
                        word_stream = self._opts.word_tokenizer.stream()
                        self._segments_ch.send_nowait(word_stream)

                    word_stream.push_text(input)
                elif isinstance(input, self._FlushSentinel):
                    if word_stream is not None:
                        word_stream.end_input()

                    word_stream = None

            self._segments_ch.close()

        @utils.log_exceptions(logger=logger)
        async def _run():
            async for word_stream in self._segments_ch:
                print('ttsV2====>word_stream e1=>', word_stream)
                await self._run_ws(word_stream)
                # await self._run_http(word_stream)
                print('ttsV2====>word_stream e2=>', word_stream)

        tasks = [
            asyncio.create_task(_tokenize_input()),
            asyncio.create_task(_run()),
        ]
        try:
            await asyncio.gather(*tasks)
        finally:
            await utils.aio.gracefully_cancel(*tasks)

    async def _run_http(
        self,
        word_stream: tokenize.WordStream,
        max_retry: int = 3,
    ) -> None:
        
        # cosy_tts_list = ['happy','angry', 'sad', 'fearful', 'surprised']
        xxxURL = _synthesize_url(None)
        # allText = 
        tmpText = ''
        flag11 = 0
        voiceType = 'happy'
        async for data in word_stream:
            token = data.token
            print(f'ttsV2 send task {data}')
            if token == "":
                continue  # empty token is closing the stream 
            tmpText += token
            print(f'xxxxxxxxxxxxxxtoken=============tts==============>{token}     tmpText:{tmpText}')
            
            if not flag11 and '生成风格:' in tmpText and '播报内容:' in tmpText:
                flag11 = 1
                # 提取生成风格
                match = re.search(r'生成风格:\s*(\w+)', tmpText)  
                if match:  
                    # \w+ 匹配一个或多个单词字符，这里就是我们想要提取的Neutral  
                    voiceType = match.group(1).lower()
                    print(f'风格======>{voiceType}')  # 输出: Neutral  
                else:  
                    print("未找到匹配项")
                # voiceType = extracted_word
                # tmpText = ''
                continue
            if not tmpText.startswith('生成风格'):
                flag11 = 1
                
            if flag11 == 0:
                continue
            # 替换 非法标签 <|HAPPY|>
            if '<|HAPPY|>' in token or 'happy' in voiceType or 'joy' in voiceType:
                # voiceType = 'happy'
                voiceType = EmotionalTTS['happy']
                # voiceType = EmotionalTTS['joy']
            elif '<|ANGRY|>' in token or 'angry' in voiceType or 'anger' in voiceType:
                voiceType = EmotionalTTS['angry']
                # voiceType = EmotionalTTS['anger']
            elif '<|SAD|>' in token or 'sad' in voiceType or 'sadness' in voiceType:
                voiceType = EmotionalTTS['sad']
                # voiceType = EmotionalTTS['sadness']
            elif '<|FEARFUL|>' in token or 'fearful' in voiceType or 'fear' in voiceType:
                # voiceType = EmotionalTTS['fearful']
                voiceType = EmotionalTTS['fear']
            elif '<|SURPRISED|>' in token or 'surprised' in voiceType or 'surprise' in voiceType:
                # voiceType = EmotionalTTS['surprised']
                voiceType = EmotionalTTS['surprise']
            elif '<|confused|>' in token or 'confused' in voiceType:
                # voiceType = EmotionalTTS['surprised']
                voiceType = EmotionalTTS['confused']
            elif voiceType in EmotionalTTS:
                voiceType = EmotionalTTS[voiceType]
            else:
                voiceType = 'happy'
            # voiceType = EmotionalTTS[voiceType]
            # 删除 非法标签:<|NEUTRAL|> <|EMO_UNKNOWN|>
            token = re.sub(r'<\|.*?\|>', '', token)
            
            if not re.search(r'<strong>.*?</strong>', token) and not re.search(r'<laughter>.*?</laughter>', token):
                # 删除所有带 <> 的标签
                token = re.sub(r'<.*?>', '', token)

            print(f'tts 合成中=================>{token}  {voiceType} {EmotionalTTS}')
            request_id = utils.shortuuid()
            segment_id = utils.shortuuid()
            parameters = {
                    "spk_id": voiceType,
                    # "instruct_text": voiceType,
                    "speed":1.0,
                    # "audio_sr":32000
                }
            inputJson = dict(
                text=token,
                model_id=self._opts.model_id,
                type=4,# 4:演示特殊定制的一个参数
                globalId="livekit-agents-test",
                parameters=parameters,
                stream=True
            )
            async with self._session.post(
                xxxURL,
                headers={},
                json=inputJson,
            ) as resp:
                async for line  in readline_from_stream(resp.content): 
                    if line:
                        result = json.loads(line)
                        if result['code'] == 200 and result.get('index', 0) != -1:
                            pcm_base64 = result['tts_speech']
                            pcmData = base64.decodebytes(pcm_base64.encode())
                            print('======ttsV2=========>', result['index'], '  pcmlen:', len(pcmData))
                            chunk_frame = rtc.AudioFrame(
                                            data=pcmData,
                                            sample_rate=self._opts.sample_rate,
                                            num_channels=1,
                                            samples_per_channel=len(pcmData) // 2,
                                )

                            self._event_ch.send_nowait(
                                tts.SynthesizedAudio(
                                    request_id=request_id,
                                    segment_id=segment_id,
                                    frame=chunk_frame,
                                )
                            )
                        elif result['index'] == -1:
                            print(result)
                            break
                        else:
                            print(f"Error: {result['message']}")
                            break
        pass
    async def _run_ws(
        self,
        word_stream: tokenize.WordStream,
        max_retry: int = 3,
    ) -> None:
        ws_conn: aiohttp.ClientWebSocketResponse | None = None
        for try_i in range(max_retry):
            retry_delay = 5
            try:
                if try_i > 0:
                    await asyncio.sleep(retry_delay)

                ws_conn = await self._session.ws_connect(
                    'ws://172.16.1.88:18101/ws_tts_audio',
                    # headers={AUTHORIZATION_HEADER: self._opts.api_key},
                )
                break
            except Exception as e:
                logger.warning(
                    f"failed to connect to 11labs, retrying in {retry_delay}s",
                    exc_info=e,
                )

        if ws_conn is None:
            raise Exception(f"failed to connect to 11labs after {max_retry} retries")

        request_id = utils.shortuuid()
        segment_id = utils.shortuuid()

        # 11labs protocol expects the first message to be an "init msg"
        init_pkt = dict(
            text=" ",
            try_trigger_generation=True,
            voice_settings=_strip_nones(dataclasses.asdict(self._opts.voice.settings))
            if self._opts.voice.settings
            else None,
            generation_config=dict(
                chunk_length_schedule=self._opts.chunk_length_schedule
            ),
        )
        # await ws_conn.send_str(json.dumps(init_pkt))
        eos_sent = False

        async def send_task():
            nonlocal eos_sent
            VoiceType = 'happy'
            parameters = {
                    "spk_id": VoiceType,
                    # "instruct_text": voiceType,
                    "speed":0.9,
                    # "audio_sr":32000
                }
            inputJson = dict(
                text='token',
                model_id=self._opts.model_id,
                type=4, # 4:演示特殊定制的一个参数
                globalId="livekit-agents-test",
                parameters=parameters,
                stream=True
            )
            async for data in word_stream:
                token = data.token
                if token == "":
                    continue  # empty token is closing the stream 
                inputJson['text'] = token
                print(f'ttsV2 send task {data}')
                await ws_conn.send_str(json.dumps(inputJson))

            # no more token, mark eos
            # eos_pkt = dict(text="")
            # await ws_conn.send_str(json.dumps(eos_pkt))
            eos_sent = True

        async def recv_task():
            nonlocal eos_sent

            while True:
                msg = await ws_conn.receive()
                # print(f'ttsV2 recv task {type(msg)}  ==> {msg}')
                if msg.type in (
                    aiohttp.WSMsgType.CLOSED,
                    aiohttp.WSMsgType.CLOSE,
                    aiohttp.WSMsgType.CLOSING,
                ):
                    if not eos_sent:
                        raise Exception(
                            "11labs connection closed unexpectedly, not all tokens have been consumed"
                        )
                    return

                if msg.type != aiohttp.WSMsgType.TEXT:
                    logger.warning("unexpected 11labs message type %s", msg.type)
                    continue

                self._process_stream_event(
                    data=json.loads(msg.data),
                    request_id=request_id,
                    segment_id=segment_id,
                )

        tasks = [
            asyncio.create_task(send_task()),
            asyncio.create_task(recv_task()),
        ]

        try:
            await asyncio.gather(*tasks)
        finally:
            await utils.aio.gracefully_cancel(*tasks)

    def _process_stream_event(
        self, *, data: dict, request_id: str, segment_id: str
    ) -> None:
        
        if data['code'] == 200 and data.get('index', 0) != -1:
            pcm_base64 = data['tts_speech']
            pcmData = base64.decodebytes(pcm_base64.encode())
            print('======ttsV2=========>', data['index'], '  pcmlen:', len(pcmData))
            chunk_frame = rtc.AudioFrame(
                        data=pcmData,
                        sample_rate=self._opts.sample_rate,
                        num_channels=1,
                        samples_per_channel=len(pcmData) // 2,
            )
            self._event_ch.send_nowait(
                tts.SynthesizedAudio(
                    request_id=request_id,
                    segment_id=segment_id,
                    frame=chunk_frame,
                )
            )
            pass
        elif data['index'] == -1:
            print(data)
            return
        else:
            print(f"Error: {data['message']}")
            return
        # encoding = _encoding_from_format(self._opts.encoding)
        # if data.get("audio"):
        #     b64data = base64.b64decode(data["audio"])
        #     if encoding == "mp3":
        #         for frame in self._mp3_decoder.decode_chunk(b64data):
        #             self._event_ch.send_nowait(
        #                 tts.SynthesizedAudio(
        #                     request_id=request_id,
        #                     segment_id=segment_id,
        #                     frame=frame,
        #                 )
        #             )
        #     else:
        #         chunk_frame = rtc.AudioFrame(
        #             data=b64data,
        #             sample_rate=self._opts.sample_rate,
        #             num_channels=1,
        #             samples_per_channel=len(b64data) // 2,
        #         )
        #         self._event_ch.send_nowait(
        #             tts.SynthesizedAudio(
        #                 request_id=request_id,
        #                 segment_id=segment_id,
        #                 frame=chunk_frame,
        #             )
        #         )
        # elif data.get("error"):
        #     logger.error("11labs reported an error: %s", data["error"])
        # elif not data.get("isFinal"):
        #     logger.error("unexpected 11labs message %s", data)
        pass


def _dict_to_voices_list(data: dict[str, Any]):
    voices: List[Voice] = []
    for voice in data["voices"]:
        voices.append(
            Voice(
                id=voice["voice_id"],
                name=voice["name"],
                category=voice["category"],
                settings=None,
            )
        )
    return voices


def _strip_nones(data: dict[str, Any]):
    return {k: v for k, v in data.items() if v is not None}


def _synthesize_url(opts: _TTSOptions) -> str:
    # base_url = opts.base_url
    # voice_id = opts.voice.id
    # model_id = opts.model_id
    # output_format = opts.encoding
    # latency = opts.streaming_latency
    # return (
    #     f"{base_url}/text-to-speech/{voice_id}/stream?"
    #     f"model_id={model_id}&output_format={output_format}&optimize_streaming_latency={latency}"
    # )
    url = API_BASE_URL_V1
    return url


def _stream_url(opts: _TTSOptions) -> str:
    base_url = opts.base_url
    voice_id = opts.voice.id
    model_id = opts.model_id
    output_format = opts.encoding
    latency = opts.streaming_latency
    return (
        f"{base_url}/text-to-speech/{voice_id}/stream-input?"
        f"model_id={model_id}&output_format={output_format}&optimize_streaming_latency={latency}"
    )
