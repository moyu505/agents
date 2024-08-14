from openpyxl import load_workbook  
  
# 加载xlsx文件  
workbook = load_workbook(filename='/mnt1/qmw/github/agents/my_plugins/turing_tts/情感控制-JZQ-240801.xlsx')  
  
# 选择工作表  
sheet = workbook['情绪控制']  # 根据你的工作表名称进行更改  
  
# 创建一个空字典来存储映射关系  
ac_mapping = {}  
  
# 遍历行（假设第一行是标题行，我们从第二行开始读取数据）  
for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始读取数据  
    # row是一个元组，包含了当前行的所有值  
    # 假设A列是第一列，C列是第三列（索引从0开始）  
    a_value = row[0]  
    c_value = row[2]  
    if a_value is None or c_value is None:  
        continue
      
    # 将A列的值作为键，C列的值作为值，添加到字典中  
    ac_mapping[a_value] = c_value  
  
# 打印字典查看结果  
print(ac_mapping)
print(ac_mapping['casual'])